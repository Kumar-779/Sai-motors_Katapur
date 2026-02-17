from flask import Flask, request, redirect, session, send_from_directory
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os, uuid

app = Flask(__name__)
app.secret_key = "sai_secure"

ADMIN_USER = "saimotors"
ADMIN_PASS = "907168"

FILE = "bikes.xlsx"
IMG_DIR = "bike_images"
os.makedirs(IMG_DIR, exist_ok=True)

# ---------- CREATE EXCEL ----------
if not os.path.exists(FILE):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "BikeNo","Brand","Model","Year","BuyPrice",
        "PurchaseTime","SellPrice","SellTime","Profit",
        "Status","Seller","Phone",
        "Img1","Img2","Img3","Img4"
    ])
    wb.save(FILE)

# ================= LOGIN =================
@app.route("/", methods=["GET","POST"])
def login():
    msg=""
    if request.method=="POST":
        if request.form["u"]==ADMIN_USER and request.form["p"]==ADMIN_PASS:
            session["login"]=True
            return redirect("/dashboard")
        else:
            msg="Wrong Login"

    return f"""
    <style>
    body{{font-family:Arial;background:#2874f0;text-align:center;color:white}}
    .box{{background:white;color:black;width:300px;margin:80px auto;padding:20px;border-radius:8px}}
    </style>
    <div class='box'>
    <h2>SAI MOTORS</h2>
    <form method=post>
    <input name=u placeholder='Username'><br><br>
    <input name=p type=password placeholder='Password'><br><br>
    <button>Login</button>
    </form>
    <p style='color:red'>{msg}</p>
    </div>
    """

# ================= DASHBOARD =================
@app.route("/dashboard", methods=["GET","POST"])
def dashboard():
    if not session.get("login"):
        return redirect("/")

    wb = load_workbook(FILE)
    ws = wb.active

    # ---------- ADD BIKE ----------
    if request.method=="POST" and "add" in request.form:
        imgs=[]
        for i in range(1,5):
            f=request.files[f"img{i}"]
            unique = str(uuid.uuid4()) + "_" + f.filename
            f.save(os.path.join(IMG_DIR,unique))
            imgs.append(unique)

        ws.append([
            request.form["no"],
            request.form["brand"],
            request.form["model"],
            request.form["year"],
            int(request.form["buy"]),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            0,"-",0,"Available","","",
            imgs[0],imgs[1],imgs[2],imgs[3]
        ])
        wb.save(FILE)
        return redirect("/dashboard")

    # ---------- SELL BIKE ----------
    if request.method=="POST" and "sell" in request.form:
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value)==request.form["sell_no"] and row[9].value=="Available":
                sell=int(request.form["sell_price"])
                buy=int(row[4].value)
                row[6].value=sell
                row[7].value=datetime.now().strftime("%Y-%m-%d %H:%M")
                row[8].value=sell-buy
                row[9].value="Sold"
                row[10].value=request.form["seller"]
                row[11].value=request.form["phone"]
        wb.save(FILE)
        return redirect("/dashboard")

    # ---------- DELETE ----------
    if request.method=="POST" and "delete" in request.form:
        bike_no=request.form["del_no"]
        for i,row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value)==bike_no:
                ws.delete_rows(i)
                break
        wb.save(FILE)
        return redirect("/dashboard")

    bikes=[row for row in ws.iter_rows(min_row=2, values_only=True)]

    # ---------- TOTALS ----------
    total_purchase = sum(int(b[4]) for b in bikes if b[4])
    total_sell = sum(int(b[6]) for b in bikes if b[9]=="Sold" and b[6])
    total_profit = sum(int(b[8]) for b in bikes if b[9]=="Sold")

    search=request.args.get("search","")
    selected=None
    found=False
    if search:
        for b in bikes:
            if str(b[0])==search:
                selected=b
                found=True

    # ================= UI =================
    html=f"""
    <h2 style='background:#2874f0;color:white;padding:10px'>SAI MOTORS KATAPUR</h2>

    <div style='display:flex;justify-content:center;gap:15px;flex-wrap:wrap;margin:10px'>
    <div style='background:white;padding:10px 20px;border-radius:8px;box-shadow:0 0 5px #ccc'>
    Total Purchase<br><b>₹ {total_purchase}</b></div>

    <div style='background:white;padding:10px 20px;border-radius:8px;box-shadow:0 0 5px #ccc'>
    Total Sell<br><b>₹ {total_sell}</b></div>

    <div style='background:white;padding:10px 20px;border-radius:8px;box-shadow:0 0 5px #ccc'>
    Total Profit<br><b>₹ {total_profit}</b></div>
    </div>

    <form method=get>
    <input name=search placeholder='Search Bike Number'>
    <button>Search</button>
    </form>
    """

    if search and not found:
        html+="<h3 style='color:red'>Bike Not Available ❌</h3>"

    # ---------- TABLE ----------
    html+="""<table>
    <tr>
    <th>Bike</th><th>Buy</th><th>Sell</th><th>Profit</th><th>Status</th><th>Delete</th>
    </tr>"""

    for b in bikes:
        color = "#c8f7c5" if b[9]=="Sold" else "#ffcccc"
        html+=f"""
        <tr style='background:{color}'>
        <td>{b[0]}</td>
        <td>{b[4]}</td>
        <td>{b[6]}</td>
        <td>{b[8]}</td>
        <td>{b[9]}</td>
        <td>
        <form method=post>
        <input type=hidden name=del_no value='{b[0]}'>
        <button name=delete>Delete</button>
        </form>
        </td>
        </tr>
        """
    html+="</table>"

    # ---------- DETAILS ----------
    if selected:
        imgs = "".join([f"<a href='/img/{selected[i]}' target=_blank><img src='/img/{selected[i]}' width=90></a>" for i in range(12,16)])

        if selected[9]=="Available":
            status="<h3 style='color:red'>AVAILABLE 🔴</h3>"
            sell_box=f"""
            <form method=post>
            <input type=hidden name=sell_no value='{selected[0]}'>
            Sell Price:<input name=sell_price required><br>
            Seller:<input name=seller required><br>
            Phone:<input name=phone required><br>
            <button name=sell>Sell Now</button>
            </form>
            """
        else:
            status=f"""
            <h3 style='color:green'>SOLD 🟢</h3>
            Seller: {selected[10]}<br>
            Phone: {selected[11]}<br>
            """
            sell_box=""

        html+=f"""
        <div class='card'>
        <h3>Bike Details</h3>
        Bike: {selected[0]}<br>
        Buy: {selected[4]}<br>
        Purchase: {selected[5]}<br>
        Sell: {selected[6]}<br>
        Profit: {selected[8]}<br>
        {status}
        {imgs}
        {sell_box}
        </div>
        """

    # ---------- ADD BIKE ----------
    html+="""
    <hr>
    <h3>Add Bike</h3>
    <form method=post enctype=multipart/form-data>
    Bike No:<input name=no required>
    Brand:<input name=brand required>
    Model:<input name=model required>
    Year:<input name=year required><br>
    Buy Price:<input name=buy required><br>
    Img1:<input type=file name=img1 required>
    Img2:<input type=file name=img2 required><br>
    Img3:<input type=file name=img3 required>
    Img4:<input type=file name=img4 required><br>
    <button name=add>Add Bike</button>
    </form>
    """

    return f"""
    <style>
    body{{font-family:Arial;background:#f1f3f6;text-align:center}}
    table{{width:90%;margin:auto;background:white;border-collapse:collapse}}
    th,td{{padding:8px;border-bottom:1px solid #ddd}}
    .card{{background:white;width:350px;margin:auto;padding:15px;box-shadow:0 0 10px #ccc}}
    button{{background:#2874f0;color:white;border:none;padding:6px 10px}}
    </style>
    {html}
    """

# ---------- IMAGE ----------
@app.route("/img/<name>")
def img(name):
    return send_from_directory(IMG_DIR,name)

# ---------- RUN ----------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
